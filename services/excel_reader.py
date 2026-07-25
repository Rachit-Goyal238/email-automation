"""
excel_reader.py

Handles loading and validating the uploaded Excel workbook.
"""

from pathlib import Path
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReader:
    """
    Reads the generated audit workbook.
    """

    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook: Workbook | None = None

    def load(self) -> Workbook:
        """
        Load the workbook.
        """

        try:
            self.workbook = load_workbook(
                filename=self.file_path,
                data_only=True
            )

            return self.workbook

        except Exception as e:
            raise Exception(f"Unable to open workbook.\n{e}")

    def get_sheet(self, sheet_name: str) -> Worksheet:
        """
        Return worksheet by name.
        """

        if self.workbook is None:
            raise Exception("Workbook not loaded.")

        if sheet_name not in self.workbook.sheetnames:
            raise Exception(
                f"Worksheet '{sheet_name}' not found."
            )

        return self.workbook[sheet_name]

    def get_checklist_sheet(self) -> Worksheet:
        """
        Returns Checklist worksheet.
        """

        return self.get_sheet("Checklist")

    def get_score_sheet(self) -> Worksheet:
        """
        Returns Score Parameters worksheet.
        """

        return self.get_sheet("Score Parameters")

    def get_sheet_names(self) -> list[str]:
        """
        Return all sheet names.
        """

        if self.workbook is None:
            raise Exception("Workbook not loaded.")

        return self.workbook.sheetnames

    def validate(self) -> bool:
        """
        Validate required worksheets exist.
        """

        required = [
            "Checklist",
            "Score Parameters"
        ]

        names = self.get_sheet_names()

        missing = []

        for sheet in required:

            if sheet not in names:
                missing.append(sheet)

        if missing:
            raise Exception(
                f"Workbook missing sheet(s): {', '.join(missing)}"
            )

        return True